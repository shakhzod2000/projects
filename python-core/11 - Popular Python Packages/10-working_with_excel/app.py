import openpyxl


# wb = openpyxl.Workbook()
path = '/mnt/c/Users/ssher/OneDrive/transactions.xlsx'
wb = openpyxl.load_workbook(path)
print(wb.sheetnames)

# sheet = wb.active
sheet1 = wb['Sheet1']
# wb.create_sheet('Sheet2', 0)
# wb.create_sheet('Sheet3', 0)
# wb.remove(sheet3)
cellB2 = sheet1['b2']
cellA2 = sheet1.cell(row=2, column=1) # A2 cell
print('sheet1.max_row:', sheet1.max_row)
print('sheet1.max_column:', sheet1.max_column)
print(cellA2.value)
cellB2.value = '1001'
cellA2.value = '1'
print(cellA2.value)
print(cellA2.row)
print(cellA2.column)
print(cellA2.coordinate)

column = sheet1['a'] # A column
dfs_cells = sheet1['a:c'] # dfs (tuple of rows first)
bfs_cells = sheet1['a1:c4'] # bfs (tuple of columns first)
print('DFS:', dfs_cells)
print('BFS:', bfs_cells)
print(sheet1[1:3]) # all used cols from row 1 to 3

# sheet1.append([1, 2, 3]) # add row at the end of sheet
# sheet1.insert_rows(idx=3, amount=1) # adds empty row at the end of sheet
# sheet1.insert_cols(idx=3, amount=1) # adds empty col at the end of sheet
sheet1.delete_rows(idx=6, amount=1)
# sheet1.delete_cols(idx=3, amount=1)

# for r in range(1, sheet1.max_row + 1):
#     for c in range(1, sheet1.max_column + 1):
#         cell = sheet1.cell(r, c)
#         print(cell.value)

wb.save(path) # <-- this writes the changes
print("Workbook updated successfully.")
